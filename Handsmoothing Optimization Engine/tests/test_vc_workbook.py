from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from handsmoothing_soe.model import CurveSolution
from handsmoothing_soe.template_output import write_template_workbook
from handsmoothing_soe.wais_workbook import load_handsmoothing_workbook


def _vc_workbook(path: Path) -> None:
    workbook = Workbook()
    hand = workbook.active
    hand.title = "handsmoothed"
    freq = workbook.create_sheet("freq")
    for column in range(2, 15):
        hand.cell(1, column, f"ss{column - 1}")
        freq.cell(1, column, f"g{column - 1}")
    for row, raw_score in enumerate(range(46), start=2):
        hand.cell(row, 1, raw_score)
        freq.cell(row, 1, raw_score)
        for column in range(2, 15):
            hand.cell(row, column, min(19, 1 + raw_score * 18 // 45))
            freq.cell(row, column, 2)
    workbook.save(path)


def test_vc_loader_reads_46_scores_and_13_groups(tmp_path: Path) -> None:
    path = tmp_path / "vc.xlsx"
    _vc_workbook(path)

    source = load_handsmoothing_workbook(path)

    assert source.workbook_layout == "wais_vc"
    assert source.output_sheet == "handsmoothed"
    assert len(source.problems) == 13
    assert list(source.problems[0].raw_scores) == list(range(46))
    assert len(source.problems[0].frequencies) == 46


def test_template_output_supports_vc_grid(tmp_path: Path) -> None:
    template = tmp_path / "vc.xlsx"
    output = tmp_path / "optimized.xlsx"
    _vc_workbook(template)
    scores = [min(19, 1 + raw_score * 18 // 45) for raw_score in range(46)]
    solutions = [
        CurveSolution(f"ss{group}", "OPTIMAL", 0.1, 0.0, scores, 10.0, 3.0, 0, 0, 0.0, 0)
        for group in range(1, 14)
    ]

    write_template_workbook(template, output, solutions, sheet_name="handsmoothed")

    workbook = load_workbook(output, data_only=True)
    assert workbook["handsmoothed"]["N47"].value == 19
    workbook.close()
