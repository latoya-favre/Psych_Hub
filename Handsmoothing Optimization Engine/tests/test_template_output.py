from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

from handsmoothing_soe.model import CurveSolution
from handsmoothing_soe.template_output import write_template_workbook


def _solution(group: int) -> CurveSolution:
    scores = [1] + [min(19, 1 + round(row * 18 / 26)) for row in range(1, 26)] + [19]
    return CurveSolution(f"ss{group}", "OPTIMAL", 0.1, 0.0, scores, 10.0, 3.0, 0, 0, 0.0, 0)


def test_template_output_preserves_formula_and_conditional_formatting(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hdsmth"
    for row in range(2, 29):
        for column in range(2, 15):
            sheet.cell(row, column, 1)
    sheet["P2"] = "=B2-C2"
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    sheet.conditional_formatting.add("B2:N28", CellIsRule(operator="lessThan", formula=["1"], fill=red_fill))
    workbook.create_sheet("freq")
    workbook.create_sheet("dif")
    workbook.save(template)

    write_template_workbook(template, output, [_solution(group) for group in range(1, 14)])
    generated = load_workbook(output, data_only=False)
    assert generated["Hdsmth"]["P2"].value == "=B2-C2"
    assert generated["Hdsmth"]["N28"].value == 19
    assert len(generated["Hdsmth"].conditional_formatting) == 1
    generated.close()

    with ZipFile(template) as before, ZipFile(output) as after:
        assert before.read("xl/styles.xml") == after.read("xl/styles.xml")

