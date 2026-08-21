from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .model import CurveProblem


@dataclass(frozen=True)
class WaisWorkbookData:
    problems: list[CurveProblem]
    manual_scores: list[list[int]]
    workbook_layout: str = "wais_mr"
    output_sheet: str = "Hdsmth"


def _frequency(value: object) -> float:
    if value in (None, ".", ""):
        return 0.0
    return float(value)


def load_wais_mr_workbook(path: str | Path) -> WaisWorkbookData:
    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise FileNotFoundError(workbook_path)
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    required = {"Hdsmth", "freq", "dif"}
    missing = required.difference(workbook.sheetnames)
    if missing:
        raise ValueError(f"Workbook is missing required sheets: {', '.join(sorted(missing))}")

    hand = workbook["Hdsmth"]
    frequency_sheet = workbook["freq"]
    empirical_sheet = workbook["dif"]
    raw_scores = [int(empirical_sheet.cell(row=row, column=1).value) for row in range(2, 29)]
    manual_scores: list[list[int]] = []
    problems: list[CurveProblem] = []
    for column in range(2, 15):
        empirical = [int(empirical_sheet.cell(row=row, column=column).value) for row in range(2, 29)]
        manual = [int(hand.cell(row=row, column=column).value) for row in range(2, 29)]
        frequencies = [_frequency(frequency_sheet.cell(row=row, column=column).value) for row in range(2, 29)]
        header = str(empirical_sheet.cell(row=1, column=column).value or f"group_{column - 1}")
        problems.append(CurveProblem(header, raw_scores, empirical, frequencies))
        manual_scores.append(manual)
    workbook.close()
    return WaisWorkbookData(problems, manual_scores)


def load_wais_vc_workbook(path: str | Path) -> WaisWorkbookData:
    """Read the WAIS5 VC norm-development workbook.

    The workbook contains final hand-smoothed curves and raw-score frequencies,
    but no separate pre-smoothed scaled-score grid.  For the initial VC
    reconstruction benchmark, the hand-smoothed curves are therefore used as
    both the empirical fidelity target and the manual comparison curves.
    """
    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise FileNotFoundError(workbook_path)
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    required = {"handsmoothed", "freq"}
    missing = required.difference(workbook.sheetnames)
    if missing:
        workbook.close()
        raise ValueError(f"Workbook is missing required VC sheets: {', '.join(sorted(missing))}")

    hand = workbook["handsmoothed"]
    frequency_sheet = workbook["freq"]
    group_columns = range(2, 15)
    data_rows = range(2, 48)
    raw_scores = [int(hand.cell(row=row, column=1).value) for row in data_rows]
    frequency_raw_scores = [int(frequency_sheet.cell(row=row, column=1).value) for row in data_rows]
    if raw_scores != frequency_raw_scores:
        workbook.close()
        raise ValueError("VC raw scores differ between handsmoothed and freq sheets")

    manual_scores: list[list[int]] = []
    problems: list[CurveProblem] = []
    for column in group_columns:
        manual = [int(hand.cell(row=row, column=column).value) for row in data_rows]
        frequencies = [_frequency(frequency_sheet.cell(row=row, column=column).value) for row in data_rows]
        header = str(hand.cell(row=1, column=column).value or f"ss{column - 1}")
        problems.append(CurveProblem(header, raw_scores, manual, frequencies))
        manual_scores.append(manual)
    workbook.close()
    return WaisWorkbookData(problems, manual_scores, "wais_vc", "handsmoothed")


def load_handsmoothing_workbook(path: str | Path) -> WaisWorkbookData:
    """Detect and load one of the supported handsmoothing workbook layouts."""
    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise FileNotFoundError(workbook_path)
    workbook = load_workbook(workbook_path, read_only=True)
    sheet_names = set(workbook.sheetnames)
    workbook.close()
    if {"Hdsmth", "freq", "dif"}.issubset(sheet_names):
        return load_wais_mr_workbook(workbook_path)
    if {"handsmoothed", "freq"}.issubset(sheet_names):
        return load_wais_vc_workbook(workbook_path)
    raise ValueError(
        "Unsupported workbook layout; expected MR sheets (Hdsmth, freq, dif) "
        "or VC sheets (handsmoothed, freq)"
    )
