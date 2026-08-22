from __future__ import annotations

import argparse
from dataclasses import dataclass
from math import exp, sin
from pathlib import Path
from random import Random

from openpyxl import Workbook


RAW_SCORES = list(range(27))
GROUP_HEADERS = [
    "age_06_00",
    "age_06_04",
    "age_06_08",
    "age_07_00",
    "age_07_04",
    "age_07_08",
    "age_08_00",
    "age_08_04",
    "age_08_08",
    "age_09_00",
    "age_09_04",
    "age_09_08",
    "age_10_00",
]
RANDOM_SEED = 20260810


@dataclass(frozen=True)
class SyntheticCurve:
    name: str
    empirical_scores: list[int]
    frequencies: list[float]


def _largest_remainder(total: int, weights: list[float], cap: int) -> list[int]:
    scaled = [total * weight / sum(weights) for weight in weights]
    steps = [min(cap, int(value)) for value in scaled]
    remaining = total - sum(steps)
    order = sorted(
        range(len(weights)),
        key=lambda index: (scaled[index] - int(scaled[index]), weights[index], -index),
        reverse=True,
    )
    while remaining > 0:
        progressed = False
        for index in order:
            if steps[index] >= cap:
                continue
            steps[index] += 1
            remaining -= 1
            progressed = True
            if remaining == 0:
                break
        if not progressed:
            raise ValueError("Unable to allocate all score steps within the configured cap")
    return steps


def _build_empirical_scores(group_index: int) -> list[int]:
    midpoint = 12.5 + (group_index - 6) * 0.18
    spread = 3.7 + (group_index % 4) * 0.2
    weights = []
    for edge in range(26):
        bell = exp(-((edge - midpoint) ** 2) / (2 * spread * spread))
        ripple = 1.0 + 0.18 * sin((edge + 1) * 0.8 + group_index * 0.6)
        weights.append(max(0.05, 0.08 + bell * ripple))
    steps = _largest_remainder(18, weights, cap=3)
    scores = [1]
    for step in steps:
        scores.append(min(19, scores[-1] + step))
    scores[-1] = 19
    return scores


def _build_frequencies(group_index: int, rng: Random) -> list[float]:
    midpoint = 13.0 + (group_index - 6) * 0.35
    spread = 4.0 + (group_index % 3) * 0.35
    frequencies: list[float] = []
    for raw_score in RAW_SCORES:
        bell = exp(-((raw_score - midpoint) ** 2) / (2 * spread * spread))
        shoulder = exp(-((raw_score - (midpoint - 4.5)) ** 2) / (2 * (spread + 1.3) ** 2))
        baseline = 0.02 * (raw_score % 3)
        jitter = rng.uniform(-0.08, 0.08)
        value = baseline + 4.8 * bell + 0.8 * shoulder + jitter
        frequencies.append(round(max(0.0, value), 4))
    return frequencies


def build_curves() -> list[SyntheticCurve]:
    rng = Random(RANDOM_SEED)
    curves: list[SyntheticCurve] = []
    for group_index, name in enumerate(GROUP_HEADERS):
        curves.append(
            SyntheticCurve(
                name=name,
                empirical_scores=_build_empirical_scores(group_index),
                frequencies=_build_frequencies(group_index, rng),
            )
        )
    return curves


def write_workbook(output_path: Path) -> Path:
    curves = build_curves()
    workbook = Workbook()
    handsheet = workbook.active
    handsheet.title = "Hdsmth"
    frequency_sheet = workbook.create_sheet("freq")
    empirical_sheet = workbook.create_sheet("dif")
    notes_sheet = workbook.create_sheet("smooth")

    for sheet in (handsheet, frequency_sheet, empirical_sheet):
        sheet.freeze_panes = "B2"
        sheet.column_dimensions["A"].width = 12
        for column_index in range(2, 15):
            sheet.column_dimensions[chr(ord("A") + column_index - 1)].width = 12

    headers = ["wisc5_raw", *[curve.name for curve in curves]]
    for column_index, header in enumerate(headers, start=1):
        handsheet.cell(row=1, column=column_index, value=header)
        frequency_sheet.cell(row=1, column=column_index, value=header)
        empirical_sheet.cell(row=1, column=column_index, value=header)

    for row_index, raw_score in enumerate(RAW_SCORES, start=2):
        handsheet.cell(row=row_index, column=1, value=raw_score)
        frequency_sheet.cell(row=row_index, column=1, value=raw_score)
        empirical_sheet.cell(row=row_index, column=1, value=raw_score)
        for column_index, curve in enumerate(curves, start=2):
            handsheet.cell(row=row_index, column=column_index, value=curve.empirical_scores[row_index - 2])
            frequency_sheet.cell(row=row_index, column=column_index, value=curve.frequencies[row_index - 2])
            empirical_sheet.cell(row=row_index, column=column_index, value=curve.empirical_scores[row_index - 2])

    notes_sheet["A1"] = "Synthetic WISC5 handsmoothing practice dataset"
    notes_sheet["A2"] = "This workbook is fake data generated for solver and handsmoothing exercises."
    notes_sheet["A3"] = f"Random seed: {RANDOM_SEED}"
    notes_sheet["A5"] = "Hdsmth and dif begin with the same empirical curve so you can edit Hdsmth manually if needed."

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Create a synthetic WISC5-style handsmoothing workbook")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/WISC5/WISC5_fake_HandSmoothing.xlsx"),
        help="Where to write the generated workbook",
    )
    args = parser.parse_args()
    output_path = write_workbook(args.output)
    print(output_path.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())