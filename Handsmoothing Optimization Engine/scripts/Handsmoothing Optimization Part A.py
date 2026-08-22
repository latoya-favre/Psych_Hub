from __future__ import annotations

"""Part A: reconstruct WAIS-5 MR starting curves from CTT frequencies.

The historical pre-handsmoothed table was not located. The documented fallback
uses midpoint percentile ranks and a normal scaled-score metric (M=10, SD=3).
The final ``mr norms.xlsx`` curves are retained only as a benchmark.
"""

import argparse
from datetime import datetime, timezone
import hashlib
import json
import math
import os
from pathlib import Path
import subprocess
from statistics import NormalDist
from typing import Any

from openpyxl import load_workbook


PROJECT_DIR = Path(r"C:\Users\UFAVRLA\PythonProjects\Handsmoothing Optimization Engine")
MR_DIR = PROJECT_DIR / "data" / "WAIS5" / "MR"
CTT_WORKBOOK = MR_DIR / "CTT_for_WAIS5_stdz_18APR2024.xls"
BENCHMARK_WORKBOOK = MR_DIR / "mr norms.xlsx"
OUTPUT_DIR = PROJECT_DIR / "outputs" / "wais_mr_reconstructed"
PREP_FILENAME = "part_a_dataprep.json"
SCHEMA = "handsmoothing.part_a.v2"
SCORE_MEAN, SCORE_SD = 10.0, 3.0
SCORE_MIN, SCORE_MAX = 1, 19
PERCENTILE_EPSILON = 0.0001

CTT_OVERRIDE = os.environ.get("HANDSMOOTHING_CTT_WORKBOOK", "").strip()
BENCHMARK_OVERRIDE = os.environ.get("HANDSMOOTHING_BENCHMARK_WORKBOOK", "").strip()
OUTPUT_OVERRIDE = os.environ.get("HANDSMOOTHING_PART_A_OUTPUT_DIR", "").strip()


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _powershell_executable() -> str:
    for candidate in ("pwsh", "powershell"):
        try:
            result = subprocess.run(
                [candidate, "-NoProfile", "-Command", "$PSVersionTable.PSVersion.Major"],
                capture_output=True, text=True, timeout=10, check=False,
            )
        except (FileNotFoundError, subprocess.TimeoutExpired):
            continue
        if result.returncode == 0:
            return candidate
    raise RuntimeError("PowerShell is required to read the legacy CTT .xls workbook")


def read_ctt_mr_frequencies(path: Path) -> tuple[list[int], list[list[float]]]:
    """Read Freq_MR through installed Excel because openpyxl cannot read .xls."""
    script = r"""
$ErrorActionPreference = 'Stop'
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {
    $book = $excel.Workbooks.Open($env:HANDSMOOTHING_CTT_READ_PATH, 0, $true)
    $sheet = $book.Worksheets.Item('Freq_MR')
    $rows = @()
    for ($r = 2; $r -le 28; $r++) {
        $values = @([int]$sheet.Cells.Item($r, 1).Value2)
        for ($g = 1; $g -le 13; $g++) {
            $value = $sheet.Cells.Item($r, 2 + 3 * ($g - 1)).Value2
            if ($null -eq $value -or $value -eq '') { $values += 0.0 }
            else { $values += [double]$value }
        }
        $rows += ,$values
    }
    $rows | ConvertTo-Json -Compress
    $book.Close($false)
}
finally {
    if ($null -ne $book) { [void][Runtime.InteropServices.Marshal]::ReleaseComObject($book) }
    $excel.Quit()
    [void][Runtime.InteropServices.Marshal]::ReleaseComObject($excel)
}
"""
    environment = os.environ.copy()
    environment["HANDSMOOTHING_CTT_READ_PATH"] = str(path.resolve())
    result = subprocess.run(
        [_powershell_executable(), "-NoProfile", "-Command", script],
        capture_output=True, text=True, timeout=90, env=environment, check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(f"Could not read Freq_MR with Excel: {result.stderr.strip()}")
    rows = json.loads(result.stdout)
    if len(rows) != 27 or any(len(row) != 14 for row in rows):
        raise ValueError("Freq_MR must contain raw scores 0-26 and 13 norm groups")
    raw_scores = [int(row[0]) for row in rows]
    if raw_scores != list(range(27)):
        raise ValueError(f"Unexpected Freq_MR raw-score grid: {raw_scores}")
    groups = [[float(row[group]) for row in rows] for group in range(1, 14)]
    for group, values in enumerate(groups, start=1):
        if any(value < 0 or not math.isfinite(value) for value in values):
            raise ValueError(f"F_{group} contains invalid frequencies")
        if sum(values) <= 1:
            raise ValueError(f"F_{group} total frequency must exceed 1")
    return raw_scores, groups


def read_benchmark(path: Path) -> tuple[list[int], list[list[int]], list[list[float]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        missing = {"handsmth", "freq"}.difference(workbook.sheetnames)
        if missing:
            raise ValueError(f"Benchmark workbook is missing sheets: {', '.join(sorted(missing))}")
        hand, freq = workbook["handsmth"], workbook["freq"]
        raw_scores = [int(hand.cell(row=row, column=1).value) for row in range(2, 29)]
        if raw_scores != list(range(27)):
            raise ValueError("Benchmark handsmth raw scores must be 0-26")
        if [int(freq.cell(row=row, column=1).value) for row in range(2, 29)] != raw_scores:
            raise ValueError("Benchmark handsmth and freq raw-score columns differ")
        scores = [[int(hand.cell(row=r, column=c).value) for r in range(2, 29)] for c in range(2, 15)]
        frequencies = [[
            0.0 if freq.cell(row=r, column=c).value in (None, "", ".") else float(freq.cell(row=r, column=c).value)
            for r in range(2, 29)
        ] for c in range(2, 15)]
        return raw_scores, scores, frequencies
    finally:
        workbook.close()


def reconstruct_scores(frequencies: list[float]) -> list[int]:
    total, below = sum(frequencies), 0.0
    scores: list[int] = []
    normal = NormalDist()
    for frequency in frequencies:
        percentile = (below + frequency / 2.0) / total
        percentile = min(max(percentile, PERCENTILE_EPSILON), 1.0 - PERCENTILE_EPSILON)
        scaled = math.floor(SCORE_MEAN + SCORE_SD * normal.inv_cdf(percentile) + 0.5)
        scores.append(min(SCORE_MAX, max(SCORE_MIN, scaled)))
        below += frequency
    return scores


def build_payload(ctt_path: Path, benchmark_path: Path) -> dict[str, Any]:
    raw_scores, ctt_frequencies = read_ctt_mr_frequencies(ctt_path)
    benchmark_raw, historical_scores, benchmark_frequencies = read_benchmark(benchmark_path)
    if raw_scores != benchmark_raw:
        raise ValueError("CTT and benchmark raw-score grids differ")
    reconstructed = [reconstruct_scores(group) for group in ctt_frequencies]
    comparisons, problems = [], []
    for group in range(13):
        name = f"ss{group + 1}"
        comparisons.append({
            "name": name,
            "ctt_n": sum(ctt_frequencies[group]),
            "benchmark_n": sum(benchmark_frequencies[group]),
            "frequency_l1_difference": sum(abs(a - b) for a, b in zip(ctt_frequencies[group], benchmark_frequencies[group])),
            "initial_vs_historical_changed_cells": sum(a != b for a, b in zip(reconstructed[group], historical_scores[group])),
        })
        problems.append({
            "name": name, "raw_scores": raw_scores,
            "empirical_scores": reconstructed[group], "frequencies": ctt_frequencies[group],
        })
    return {
        "schema": SCHEMA,
        "prepared_at_utc": datetime.now(timezone.utc).isoformat(),
        "project": "WAIS-5", "subtest": "Matrix Reasoning",
        "workbook_layout": "wais_mr_reconstructed_from_ctt",
        "source_workbook": str(benchmark_path.resolve()), "output_sheet": "handsmth",
        "ctt_workbook": str(ctt_path.resolve()),
        "source_fingerprints": {"ctt_sha256": file_sha256(ctt_path), "benchmark_sha256": file_sha256(benchmark_path)},
        "empirical_score_source": "ctt_freq_mr_midpoint_percentile_normal_score_reconstruction",
        "reconstruction": {
            "percentile_rule": "(cumulative_frequency_below + 0.5 * cell_frequency) / group_n",
            "normal_score_rule": "round_half_up(10 + 3 * inverse_standard_normal(percentile))",
            "bounds": [SCORE_MIN, SCORE_MAX], "percentile_epsilon": PERCENTILE_EPSILON,
            "historical_rule_confirmed": False,
        },
        "group_count": 13, "row_count": 27, "problems": problems,
        "manual_scores": historical_scores, "historical_scores": historical_scores,
        "source_comparisons": comparisons,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Prepare WAIS-5 MR curves from CTT frequencies")
    parser.add_argument("--ctt-workbook", type=Path, help="CTT .xls containing Freq_MR")
    parser.add_argument("--benchmark-workbook", type=Path, help="Historical final mr norms.xlsx")
    parser.add_argument("--output-dir", type=Path, help="Part A output directory")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    ctt_path = args.ctt_workbook or (Path(CTT_OVERRIDE) if CTT_OVERRIDE else CTT_WORKBOOK)
    benchmark_path = args.benchmark_workbook or (Path(BENCHMARK_OVERRIDE) if BENCHMARK_OVERRIDE else BENCHMARK_WORKBOOK)
    output_dir = args.output_dir or (Path(OUTPUT_OVERRIDE) if OUTPUT_OVERRIDE else OUTPUT_DIR)
    for path in (ctt_path, benchmark_path):
        if not path.is_file():
            raise FileNotFoundError(path)
    payload = build_payload(ctt_path, benchmark_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    prep_path = output_dir / PREP_FILENAME
    prep_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print("Handsmoothing Optimization Part A — WAIS-5 MR")
    print(f"CTT input: {ctt_path}")
    print(f"Historical benchmark: {benchmark_path}")
    print("Reconstruction: midpoint percentile -> normal scaled score (M=10, SD=3)")
    for item in payload["source_comparisons"]:
        print(f"{item['name']}: CTT N={item['ctt_n']:.0f}; benchmark N={item['benchmark_n']:.0f}; initial/final changed={item['initial_vs_historical_changed_cells']}")
    print(f"Part A artifact: {prep_path}")


if __name__ == "__main__":
    main()
