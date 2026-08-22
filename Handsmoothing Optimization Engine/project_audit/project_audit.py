from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path
import tempfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
WORKBOOK = ROOT / "project_audit.xlsx"
CSV_BACKUP = ROOT / "project_audit.csv"
HEADERS = ["Timestamp", "Project", "File", "Action", "Notes"]
ACTIONS = [
    "Received",
    "Created",
    "Scored",
    "Converted to Excel",
    "Cleaned",
    "Merged",
    "Validated",
    "Analyzed",
    "Exported",
    "Delivered",
    "Archived",
    "Other",
]


def initialize(*, overwrite: bool = False) -> None:
    if WORKBOOK.exists() and not overwrite:
        return

    workbook = Workbook()
    log = workbook.active
    log.title = "Audit Log"
    log.append(HEADERS)
    log.freeze_panes = "A2"
    log.auto_filter.ref = "A1:E1"
    widths = {"A": 27, "B": 18, "C": 70, "D": 24, "E": 50}
    for column, width in widths.items():
        log.column_dimensions[column].width = width
    for cell in log[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    table = Table(displayName="ProjectAudit", ref="A1:E1")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    log.add_table(table)

    lists = workbook.create_sheet("Lists")
    lists.sheet_state = "hidden"
    for row, action in enumerate(ACTIONS, start=1):
        lists.cell(row=row, column=1, value=action)
    validation = DataValidation(
        type="list", formula1=f"=Lists!$A$1:$A${len(ACTIONS)}", allow_blank=False
    )
    log.add_data_validation(validation)
    validation.add("D2:D1048576")

    workbook.save(WORKBOOK)
    with CSV_BACKUP.open("w", newline="", encoding="utf-8-sig") as handle:
        csv.writer(handle).writerow(HEADERS)


def log_entries(project: str, files: list[str], action: str, notes: str) -> None:
    initialize()
    timestamp = datetime.now().astimezone().isoformat(timespec="seconds")
    rows = [[timestamp, project, str(Path(item).expanduser().resolve()), action, notes] for item in files]

    workbook = load_workbook(WORKBOOK)
    sheet = workbook["Audit Log"]
    for row in rows:
        sheet.append(row)
    sheet.tables["ProjectAudit"].ref = f"A1:E{sheet.max_row}"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=ROOT, delete=False) as handle:
        temporary = Path(handle.name)
    try:
        workbook.save(temporary)
        temporary.replace(WORKBOOK)
    finally:
        temporary.unlink(missing_ok=True)

    with CSV_BACKUP.open("a", newline="", encoding="utf-8-sig") as handle:
        csv.writer(handle).writerows(rows)

    for row in rows:
        print(f"Logged {row[1]} | {row[3]} | {row[2]} | {row[0]}")


def parser() -> argparse.ArgumentParser:
    result = argparse.ArgumentParser(description="Append timestamped project file actions to Excel and CSV.")
    subcommands = result.add_subparsers(dest="command", required=True)
    init = subcommands.add_parser("init", help="Create the blank audit workbook")
    init.add_argument("--overwrite", action="store_true")
    log = subcommands.add_parser("log", help="Add one audit row per file")
    log.add_argument("--project", required=True)
    log.add_argument("--action", required=True)
    log.add_argument("--notes", default="")
    log.add_argument("files", nargs="+", help="One or more file paths")
    return result


def main() -> None:
    args = parser().parse_args()
    if args.command == "init":
        initialize(overwrite=args.overwrite)
        print(f"Ready: {WORKBOOK}")
    else:
        log_entries(args.project, args.files, args.action, args.notes)


if __name__ == "__main__":
    main()
